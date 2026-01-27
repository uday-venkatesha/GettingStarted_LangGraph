import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.patches import Rectangle
import warnings
warnings.filterwarnings('ignore')

# Set random seed for reproducibility
np.random.seed(42)
random.seed(42)

# Define our business parameters
CATEGORIES = {
    'Electronics': ['Smartphones', 'Laptops', 'Headphones', 'Tablets', 'Smartwatches'],
    'Home & Kitchen': ['Cookware', 'Furniture', 'Bedding', 'Storage', 'Appliances'],
    'Fashion': ['Mens Wear', 'Womens Wear', 'Footwear', 'Accessories', 'Kids Wear'],
    'Beauty': ['Skincare', 'Makeup', 'Fragrances', 'Hair Care', 'Tools'],
    'Sports': ['Fitness Equipment', 'Outdoor Gear', 'Sportswear', 'Bikes', 'Nutrition']
}

REGIONS = ['North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East']

SUPPLIERS = {
    'TechGlobal Inc': ['Electronics'],
    'HomeComfort Ltd': ['Home & Kitchen'],
    'StyleHub Corp': ['Fashion'],
    'BeautyPro Suppliers': ['Beauty'],
    'ActiveLife Co': ['Sports'],
    'MultiSource Traders': ['Electronics', 'Fashion', 'Home & Kitchen'],
    'Global Imports LLC': ['Beauty', 'Sports'],
    'PrimeLine Distributors': ['Electronics', 'Sports']
}

WAREHOUSES = {
    'WH-NA-01': 'North America',
    'WH-NA-02': 'North America', 
    'WH-EU-01': 'Europe',
    'WH-EU-02': 'Europe',
    'WH-AP-01': 'Asia Pacific',
    'WH-AP-02': 'Asia Pacific',
    'WH-LA-01': 'Latin America',
    'WH-ME-01': 'Middle East'
}

PAYMENT_METHODS = ['Credit Card', 'Debit Card', 'PayPal', 'Bank Transfer', 'Digital Wallet']

def generate_product_data(num_records=1000):
    """Generate comprehensive e-commerce dataset"""
    
    data = []
    start_date = datetime(2024, 1, 1)
    end_date = datetime(2025, 1, 25)
    
    for i in range(num_records):
        # Basic product info
        category = random.choice(list(CATEGORIES.keys()))
        subcategory = random.choice(CATEGORIES[category])
        product_id = f"PRD-{category[:3].upper()}-{i+1000:04d}"
        
        # Pricing varies by category
        if category == 'Electronics':
            base_price = random.uniform(100, 2000)
        elif category == 'Fashion':
            base_price = random.uniform(20, 300)
        elif category == 'Beauty':
            base_price = random.uniform(10, 150)
        else:
            base_price = random.uniform(30, 500)
            
        # Supplier selection
        eligible_suppliers = [s for s, cats in SUPPLIERS.items() if category in cats]
        supplier = random.choice(eligible_suppliers)
        
        # Regional distribution
        region = random.choice(REGIONS)
        warehouse = random.choice([wh for wh, reg in WAREHOUSES.items() if reg == region])
        
        # Order details
        order_date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
        quantity = random.choices([1, 2, 3, 4, 5], weights=[50, 25, 15, 7, 3])[0]
        
        # Acceptance status - most products accepted
        acceptance_probability = 0.92 if supplier in ['TechGlobal Inc', 'StyleHub Corp'] else 0.85
        is_accepted = random.random() < acceptance_probability
        rejection_reason = None if is_accepted else random.choice([
            'Quality Issues', 'Damaged in Transit', 'Incorrect Specifications', 
            'Documentation Incomplete', 'Late Delivery'
        ])
        
        # Shipping costs vary by region and product category
        base_shipping = {'North America': 15, 'Europe': 25, 'Asia Pacific': 35, 
                        'Latin America': 30, 'Middle East': 40}
        weight_factor = {'Electronics': 1.5, 'Home & Kitchen': 2.0, 'Fashion': 0.8,
                        'Beauty': 0.5, 'Sports': 1.3}
        shipping_cost = base_shipping[region] * weight_factor[category] * quantity
        
        # Payment processing
        payment_method = random.choice(PAYMENT_METHODS)
        payment_fee_rate = {'Credit Card': 0.029, 'Debit Card': 0.022, 'PayPal': 0.034,
                           'Bank Transfer': 0.01, 'Digital Wallet': 0.025}
        
        # Calculate financials
        unit_price = base_price * random.uniform(0.9, 1.1)
        gross_revenue = unit_price * quantity
        payment_fee = gross_revenue * payment_fee_rate[payment_method]
        
        # Cost of goods (60-75% of price)
        cogs = unit_price * quantity * random.uniform(0.60, 0.75)
        
        # Net profit
        net_profit = gross_revenue - cogs - shipping_cost - payment_fee if is_accepted else 0
        
        # Delivery time varies by region
        avg_delivery_days = {'North America': 3, 'Europe': 5, 'Asia Pacific': 7,
                            'Latin America': 8, 'Middle East': 9}
        delivery_days = avg_delivery_days[region] + random.randint(-1, 3)
        
        # Customer rating (only for accepted products)
        rating = round(random.gauss(4.2, 0.8), 1) if is_accepted else None
        if rating and rating > 5: rating = 5.0
        if rating and rating < 1: rating = 1.0
        
        record = {
            'Product_ID': product_id,
            'Category': category,
            'Subcategory': subcategory,
            'Supplier': supplier,
            'Region': region,
            'Warehouse': warehouse,
            'Order_Date': order_date.strftime('%Y-%m-%d'),
            'Order_Month': order_date.strftime('%Y-%m'),
            'Quantity': quantity,
            'Unit_Price': round(unit_price, 2),
            'Gross_Revenue': round(gross_revenue, 2),
            'COGS': round(cogs, 2),
            'Shipping_Cost': round(shipping_cost, 2),
            'Payment_Method': payment_method,
            'Payment_Fee': round(payment_fee, 2),
            'Net_Profit': round(net_profit, 2),
            'Is_Accepted': 'Yes' if is_accepted else 'No',
            'Rejection_Reason': rejection_reason or 'N/A',
            'Delivery_Days': delivery_days if is_accepted else None,
            'Customer_Rating': rating
        }
        
        data.append(record)
    
    return pd.DataFrame(data)

def create_monthly_summary(df):
    """Create monthly KPI summary"""
    
    # Filter accepted orders only
    accepted = df[df['Is_Accepted'] == 'Yes'].copy()
    
    monthly = accepted.groupby('Order_Month').agg({
        'Product_ID': 'count',
        'Gross_Revenue': 'sum',
        'COGS': 'sum',
        'Shipping_Cost': 'sum',
        'Payment_Fee': 'sum',
        'Net_Profit': 'sum',
        'Quantity': 'sum',
        'Customer_Rating': 'mean',
        'Delivery_Days': 'mean'
    }).reset_index()
    
    monthly.columns = ['Month', 'Total_Orders', 'Gross_Revenue', 'COGS', 
                       'Shipping_Cost', 'Payment_Fee', 'Net_Profit', 
                       'Units_Sold', 'Avg_Rating', 'Avg_Delivery_Days']
    
    # Calculate margins and metrics
    monthly['Gross_Margin_%'] = ((monthly['Gross_Revenue'] - monthly['COGS']) / 
                                  monthly['Gross_Revenue'] * 100)
    monthly['Net_Margin_%'] = (monthly['Net_Profit'] / monthly['Gross_Revenue'] * 100)
    monthly['Shipping_Cost_%'] = (monthly['Shipping_Cost'] / monthly['Gross_Revenue'] * 100)
    monthly['AOV'] = monthly['Gross_Revenue'] / monthly['Total_Orders']
    
    return monthly.round(2)

def create_excel_report(df, monthly_df):
    """Create formatted Excel file with multiple sheets"""
    
    wb = Workbook()
    
    # Header style
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Data style
    data_alignment = Alignment(horizontal='left', vertical='center')
    number_format = '#,##0.00'
    
    # Sheet 1: Raw Transaction Data
    ws1 = wb.active
    ws1.title = 'Transaction Data'
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = data_alignment
                
                if c_idx in [10, 11, 12, 13, 15, 16]:  # Numeric columns
                    cell.number_format = number_format
    
    # Adjust column widths
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws1.column_dimensions[column].width = min(max_length + 2, 20)
    
    # Sheet 2: Monthly Summary
    ws2 = wb.create_sheet('Monthly KPIs')
    
    for r_idx, row in enumerate(dataframe_to_rows(monthly_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = data_alignment
                if c_idx >= 2:
                    cell.number_format = number_format
    
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[column].width = min(max_length + 2, 18)
    
    # Sheet 3: KPI Dashboard Template
    ws3 = wb.create_sheet('KPI Summary')
    
    # Add KPI calculations
    accepted_count = len(df[df['Is_Accepted'] == 'Yes'])
    total_count = len(df)
    acceptance_rate = (accepted_count / total_count * 100)
    
    kpi_data = [
        ['Key Performance Indicators', ''],
        ['', ''],
        ['Metric', 'Value'],
        ['Total Transactions', total_count],
        ['Accepted Orders', accepted_count],
        ['Acceptance Rate %', f'{acceptance_rate:.2f}%'],
        ['Total Revenue', f"${df[df['Is_Accepted']=='Yes']['Gross_Revenue'].sum():,.2f}"],
        ['Total Profit', f"${df[df['Is_Accepted']=='Yes']['Net_Profit'].sum():,.2f}"],
        ['Avg Order Value', f"${df[df['Is_Accepted']=='Yes']['Gross_Revenue'].mean():,.2f}"],
        ['Total Shipping Cost', f"${df[df['Is_Accepted']=='Yes']['Shipping_Cost'].sum():,.2f}"],
        ['Avg Customer Rating', f"{df[df['Is_Accepted']=='Yes']['Customer_Rating'].mean():.2f}"]
    ]
    
    for r_idx, row in enumerate(kpi_data, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif r_idx == 3:
                cell.font = header_font
                cell.fill = header_fill
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20
    
    wb.save('/home/claude/ecommerce_data.xlsx')
    print("✓ Excel file created successfully")

def create_visualizations(df, monthly_df):
    """Generate professional visualizations"""
    
    plt.style.use('seaborn-v0_8-darkgrid')
    sns.set_palette("husl")
    
    fig = plt.figure(figsize=(20, 12))
    
    # 1. Revenue Trend Over Time
    ax1 = plt.subplot(2, 3, 1)
    monthly_sorted = monthly_df.sort_values('Month')
    ax1.plot(range(len(monthly_sorted)), monthly_sorted['Gross_Revenue'], 
             marker='o', linewidth=2.5, markersize=8, color='#2E86AB')
    ax1.fill_between(range(len(monthly_sorted)), monthly_sorted['Gross_Revenue'], 
                     alpha=0.3, color='#2E86AB')
    ax1.set_title('Monthly Revenue Trend', fontsize=14, fontweight='bold', pad=15)
    ax1.set_xlabel('Month', fontsize=11)
    ax1.set_ylabel('Revenue ($)', fontsize=11)
    ax1.grid(True, alpha=0.3)
    ax1.set_xticks(range(len(monthly_sorted)))
    ax1.set_xticklabels(monthly_sorted['Month'].str[-2:], rotation=45)
    
    # 2. Category Performance
    ax2 = plt.subplot(2, 3, 2)
    accepted = df[df['Is_Accepted'] == 'Yes']
    category_revenue = accepted.groupby('Category')['Gross_Revenue'].sum().sort_values(ascending=False)
    colors = ['#E63946', '#F77F00', '#06A77D', '#118AB2', '#073B4C']
    bars = ax2.bar(range(len(category_revenue)), category_revenue.values, color=colors, alpha=0.8)
    ax2.set_title('Revenue by Category', fontsize=14, fontweight='bold', pad=15)
    ax2.set_xlabel('Category', fontsize=11)
    ax2.set_ylabel('Revenue ($)', fontsize=11)
    ax2.set_xticks(range(len(category_revenue)))
    ax2.set_xticklabels(category_revenue.index, rotation=30, ha='right')
    
    # Add value labels on bars
    for bar in bars:
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height,
                f'${height/1000:.0f}K', ha='center', va='bottom', fontsize=9)
    
    # 3. Acceptance Rate by Supplier
    ax3 = plt.subplot(2, 3, 3)
    supplier_stats = df.groupby('Supplier').agg({
        'Is_Accepted': lambda x: (x == 'Yes').sum() / len(x) * 100
    }).sort_values('Is_Accepted', ascending=False)
    
    bars = ax3.barh(range(len(supplier_stats)), supplier_stats['Is_Accepted'], 
                     color='#06A77D', alpha=0.7)
    ax3.set_title('Supplier Acceptance Rate', fontsize=14, fontweight='bold', pad=15)
    ax3.set_xlabel('Acceptance Rate (%)', fontsize=11)
    ax3.set_yticks(range(len(supplier_stats)))
    ax3.set_yticklabels([s[:20] for s in supplier_stats.index])
    ax3.set_xlim(0, 100)
    
    for i, bar in enumerate(bars):
        width = bar.get_width()
        ax3.text(width + 1, bar.get_y() + bar.get_height()/2.,
                f'{width:.1f}%', va='center', fontsize=9)
    
    # 4. Shipping Cost Analysis
    ax4 = plt.subplot(2, 3, 4)
    region_shipping = accepted.groupby('Region')['Shipping_Cost'].sum().sort_values(ascending=False)
    wedges, texts, autotexts = ax4.pie(region_shipping.values, labels=region_shipping.index,
                                        autopct='%1.1f%%', startangle=90,
                                        colors=sns.color_palette("Set2"))
    ax4.set_title('Shipping Cost Distribution by Region', fontsize=14, fontweight='bold', pad=15)
    
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_fontweight('bold')
        autotext.set_fontsize(9)
    
    # 5. Profit Margins Over Time
    ax5 = plt.subplot(2, 3, 5)
    ax5.plot(range(len(monthly_sorted)), monthly_sorted['Gross_Margin_%'], 
             marker='s', label='Gross Margin', linewidth=2, markersize=7)
    ax5.plot(range(len(monthly_sorted)), monthly_sorted['Net_Margin_%'], 
             marker='^', label='Net Margin', linewidth=2, markersize=7)
    ax5.set_title('Margin Trends', fontsize=14, fontweight='bold', pad=15)
    ax5.set_xlabel('Month', fontsize=11)
    ax5.set_ylabel('Margin (%)', fontsize=11)
    ax5.legend(loc='best')
    ax5.grid(True, alpha=0.3)
    ax5.set_xticks(range(len(monthly_sorted)))
    ax5.set_xticklabels(monthly_sorted['Month'].str[-2:], rotation=45)
    
    # 6. Payment Method Distribution
    ax6 = plt.subplot(2, 3, 6)
    payment_counts = accepted['Payment_Method'].value_counts()
    bars = ax6.bar(range(len(payment_counts)), payment_counts.values, 
                   color=sns.color_palette("viridis", len(payment_counts)), alpha=0.8)
    ax6.set_title('Orders by Payment Method', fontsize=14, fontweight='bold', pad=15)
    ax6.set_xlabel('Payment Method', fontsize=11)
    ax6.set_ylabel('Number of Orders', fontsize=11)
    ax6.set_xticks(range(len(payment_counts)))
    ax6.set_xticklabels(payment_counts.index, rotation=30, ha='right')
    
    for bar in bars:
        height = bar.get_height()
        ax6.text(bar.get_x() + bar.get_width()/2., height,
                f'{int(height)}', ha='center', va='bottom', fontsize=9)
    
    plt.tight_layout(pad=3.0)
    plt.savefig('/home/claude/ecommerce_dashboard.png', dpi=300, bbox_inches='tight')
    print("✓ Main dashboard saved")
    
    # Create additional KPI-focused chart
    fig2, axes = plt.subplots(2, 2, figsize=(16, 10))
    
    # KPI 1: Daily order volume
    ax = axes[0, 0]
    df_accepted = df[df['Is_Accepted'] == 'Yes'].copy()
    df_accepted['Order_Date_dt'] = pd.to_datetime(df_accepted['Order_Date'])
    daily_orders = df_accepted.groupby('Order_Date_dt').size()
    ax.plot(daily_orders.index, daily_orders.values, linewidth=1.5, alpha=0.7, color='#E63946')
    ax.set_title('Daily Order Volume', fontsize=13, fontweight='bold')
    ax.set_xlabel('Date', fontsize=10)
    ax.set_ylabel('Orders', fontsize=10)
    ax.grid(True, alpha=0.3)
    
    # KPI 2: Average order value by category
    ax = axes[0, 1]
    category_aov = accepted.groupby('Category')['Gross_Revenue'].mean().sort_values(ascending=False)
    ax.barh(range(len(category_aov)), category_aov.values, color='#118AB2', alpha=0.8)
    ax.set_title('Average Order Value by Category', fontsize=13, fontweight='bold')
    ax.set_xlabel('AOV ($)', fontsize=10)
    ax.set_yticks(range(len(category_aov)))
    ax.set_yticklabels(category_aov.index)
    
    # KPI 3: Customer ratings distribution
    ax = axes[1, 0]
    ratings = accepted['Customer_Rating'].dropna()
    ax.hist(ratings, bins=20, color='#06A77D', alpha=0.7, edgecolor='black')
    ax.axvline(ratings.mean(), color='red', linestyle='--', linewidth=2, label=f'Mean: {ratings.mean():.2f}')
    ax.set_title('Customer Rating Distribution', fontsize=13, fontweight='bold')
    ax.set_xlabel('Rating', fontsize=10)
    ax.set_ylabel('Frequency', fontsize=10)
    ax.legend()
    
    # KPI 4: Regional performance matrix
    ax = axes[1, 1]
    region_metrics = accepted.groupby('Region').agg({
        'Gross_Revenue': 'sum',
        'Net_Profit': 'sum'
    })
    x = region_metrics['Gross_Revenue'] / 1000
    y = region_metrics['Net_Profit'] / 1000
    sizes = accepted.groupby('Region').size() * 3
    
    scatter = ax.scatter(x, y, s=sizes, alpha=0.6, c=range(len(x)), cmap='coolwarm', edgecolors='black')
    
    for i, region in enumerate(region_metrics.index):
        ax.annotate(region, (x.iloc[i], y.iloc[i]), fontsize=9, ha='center')
    
    ax.set_title('Regional Performance (Revenue vs Profit)', fontsize=13, fontweight='bold')
    ax.set_xlabel('Revenue ($K)', fontsize=10)
    ax.set_ylabel('Profit ($K)', fontsize=10)
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig('/home/claude/kpi_deep_dive.png', dpi=300, bbox_inches='tight')
    print("✓ KPI deep dive chart saved")

# Main execution
print("Generating e-commerce data...")
df = generate_product_data(1000)
print(f"✓ Generated {len(df)} transaction records")

print("\nCalculating monthly KPIs...")
monthly_summary = create_monthly_summary(df)
print(f"✓ Monthly summary created for {len(monthly_summary)} months")

print("\nCreating Excel report...")
create_excel_report(df, monthly_summary)

print("\nGenerating visualizations...")
create_visualizations(df, monthly_summary)

# Print summary statistics
print("\n" + "="*60)
print("DATA SUMMARY")
print("="*60)
print(f"Total Transactions: {len(df)}")
print(f"Accepted Orders: {len(df[df['Is_Accepted']=='Yes'])}")
print(f"Acceptance Rate: {len(df[df['Is_Accepted']=='Yes'])/len(df)*100:.2f}%")
print(f"Total Revenue: ${df[df['Is_Accepted']=='Yes']['Gross_Revenue'].sum():,.2f}")
print(f"Total Profit: ${df[df['Is_Accepted']=='Yes']['Net_Profit'].sum():,.2f}")
print(f"Date Range: {df['Order_Date'].min()} to {df['Order_Date'].max()}")
print("="*60)